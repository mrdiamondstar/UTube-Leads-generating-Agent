"""Excel (.xlsx) export of scored leads.

Builds a formatted workbook from (Channel, LeadScore) pairs. Used by the
`/api/v1/leads/export` endpoint and any offline export script.
"""
from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.common.countries import country_name
from app.domain.models import Channel, LeadScore, Video

# (header, width, extractor)
_COLUMNS: list[tuple[str, int, str]] = [
    ("Channel", 34, "title"),
    ("Channel URL", 42, "youtube_url"),
    ("Category", 12, "category"),
    ("Country", 18, "country"),
    ("Country Code", 12, "country_code"),
    ("Subscribers", 13, "subscriber_count"),
    ("Total Views", 15, "view_count"),
    ("Videos", 9, "video_count"),
    ("Lead Score", 11, "score"),
    ("Confidence", 11, "confidence"),
    ("Lead Category", 14, "lead_category"),
    ("Underperforming", 15, "is_underperforming"),
    ("Perf. Ratio", 11, "perf_ratio"),
    ("Latest Video", 42, "last_video_title"),
    ("Latest Video Views", 16, "last_video_views"),
    ("Latest Video URL", 42, "last_video_url"),
    ("Latest Video Posted", 20, "last_video_posted"),
    ("Website", 30, "website"),
    ("Public Email", 28, "public_email"),
    ("Socials", 40, "socials"),
    ("AI Reasoning", 60, "reasoning"),
]


def _relative_age(dt: datetime | None) -> str:
    if dt is None:
        return ""
    now = datetime.now(timezone.utc)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    secs = max(0, int((now - dt).total_seconds()))
    for name, span in (
        ("year", 31536000),
        ("month", 2592000),
        ("week", 604800),
        ("day", 86400),
        ("hour", 3600),
        ("minute", 60),
    ):
        val = secs // span
        if val >= 1:
            return f"{val} {name}{'s' if val > 1 else ''} ago"
    return "just now"

_CATEGORY_FILL = {
    "hot": "F8CBAD",
    "warm": "FFE699",
    "cold": "BDD7EE",
    "disqualified": "D9D9D9",
}


def _row_values(channel: Channel, score: LeadScore, video: Video | None) -> dict:
    contrib = score.feature_contributions or {}
    gap = contrib.get("opportunity_gap", {}).get("strength")
    perf_ratio = round(1 - gap, 3) if gap is not None else None
    socials = ", ".join(f"{k}: {v}" for k, v in (channel.social_links or {}).items())
    return {
        "title": channel.title,
        "youtube_url": f"https://www.youtube.com/channel/{channel.youtube_id}",
        "category": channel.category or "",
        "country": country_name(channel.country) or "",
        "country_code": channel.country or "",
        "subscriber_count": channel.subscriber_count,
        "view_count": channel.view_count,
        "video_count": channel.video_count,
        "score": round(score.score, 2),
        "confidence": round(score.confidence, 3),
        "lead_category": score.category,
        "is_underperforming": "Yes" if score.is_underperforming else "No",
        "perf_ratio": perf_ratio,
        "last_video_title": video.title if video else "",
        "last_video_views": video.view_count if video else "",
        "last_video_url": (
            f"https://www.youtube.com/watch?v={video.youtube_video_id}" if video else ""
        ),
        "last_video_posted": _relative_age(video.published_at) if video else "",
        "website": channel.website or "",
        "public_email": channel.public_email or "",
        "socials": socials,
        "reasoning": score.reasoning or "",
    }


def build_leads_workbook(
    pairs: Iterable[tuple[Channel, LeadScore]],
    latest_videos: dict[str, Video] | None = None,
) -> bytes:
    latest_videos = latest_videos or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")

    # Header row
    for col_idx, (header, width, _) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for r, (channel, score) in enumerate(pairs, start=2):
        values = _row_values(channel, score, latest_videos.get(channel.id))
        fill_hex = _CATEGORY_FILL.get(score.category)
        for col_idx, (_, _, key) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=r, column=col_idx, value=values[key])
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=key in ("reasoning", "socials", "last_video_title"),
            )
            if fill_hex and key == "lead_category":
                cell.fill = PatternFill("solid", fgColor=fill_hex)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNS))}{max(ws.max_row, 1)}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
